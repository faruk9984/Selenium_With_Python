

# ===========================================================
# Selenium with Python Tutorial 26-Data Driven Testing using Microsoft Excel + OpenPyXL Module
# connect with datadriventestcase.py

import openpyxl

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getColumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def readData(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetname,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)

