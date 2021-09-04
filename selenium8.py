"""
# Selenium with Python Tutorial 14-How to handle Browser Windows | Switch between the Windows

from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome(executable_path="chromedriver.exe")

driver.get("http://demo.automationtesting.in/Windows.html")

driver.find_element_by_xpath("//*[@id='Tabbed']/a/button").click()
print(driver.current_window_handle)

handles=driver.window_handles
for hand in handles:
    driver.switch_to.window(hand)
    print(driver.title)
    if driver.title=="Frames & windows":
        driver.close()

driver.quit()



================================================================
# Selenium with Python Tutorial 15- Working with HTML/Web Table

from selenium import webdriver

driver=webdriver.Chrome(executable_path="chromedriver.exe")

# driver.get("file:///c:/SeleniumPractice/sample.html")
driver.get("https://www.tutorialspoint.com/html/html_tables.htm")

rows=len(driver.find_elements_by_xpath("/html/body/table/tbody/tr"))
cols=len(driver.find_elements_by_xpath("/html/body/table/tbody/tr[1]/th"))
print(rows)
print(cols)

print("product"+"  "+"Artical"+"  "+"price")

for r in range(2,rows+1):
    for c in range(1,cols+1):
        value=driver.find_element_by_xpath("/html/body/table/tbody/tr["+str(r)+"]/td["+str(c)+"]").text
        print(value,end="   ")
    print()




# =================================================================
# Selenium with Python Tutorial 16- How to Scroll Web Pages in Selenium

from selenium import webdriver

driver=webdriver.Chrome(executable_path="chromedriver.exe")
driver.get("https://www.countries-ofthe-world.com/flags-of-the-world.html")

driver.maximize_window()
# 1 way index position
# driver.execute_script("window.scrollBy(0,1000)","")

# 2nd way specific position
flag=driver.find_element_by_xpath("//*[@id='content']/div[2]/div[2]/table[1]/tbody/tr[86]/td[1]/img")
driver.execute_script("arguments[0].scrollIntoView() ;",flag)

# 3rd way end of the page
driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")



# ==========================================================================
# Selenium with Python Tutorial 17-Handle Mouse Actions | Mouse Hover Action

from selenium import webdriver
from selenium.webdriver import ActionChains

driver=webdriver.Chrome(executable_path="chromedriver.exe")
driver.get("https://opensource-demo.orangehrmlive.com")
driver.maximize_window()

driver.find_element_by_xpath("//*[@id='txtUsername']").send_keys("Admin")
driver.find_element_by_xpath("//*[@id='txtPassword']").send_keys("admin123")
driver.find_element_by_xpath("//*[@id='btnLogin']").click()

admin=driver.find_element_by_xpath("//*[@id='menu_admin_viewAdminModule']/b")
userman=driver.find_element_by_xpath("//*[@id='menu_admin_UserManagement']")
user=driver.find_element_by_xpath("//*[@id='menu_admin_viewSystemUsers']")

actions=ActionChains(driver)
actions.move_to_element(admin).move_to_element(userman).move_to_element(user).click().perform()




# ===================================================================
# Selenium with Python Tutorial 18-Handle Mouse Actions | Double click Action

from selenium import webdriver
from selenium.webdriver import ActionChains

driver=webdriver.Chrome(executable_path="chromedriver.exe")
driver.get("http://testautomationpractice.blogspot.com/")
driver.maximize_window()

element=driver.find_element_by_xpath("//*[@id='HTML10']/div[1]/button")
actions=ActionChains(driver)
actions.double_click(element).perform()



# ===========================================================
# Selenium with Python Tutorial 19-Handle Mouse Actions | Right click Action


from selenium import webdriver
from selenium.webdriver import ActionChains

driver=webdriver.Chrome(executable_path="chromedriver.exe")
driver.get("https://swisnl.github.io/jQuery-contextMenu/demo.html")
driver.maximize_window()

right_click=driver.find_element_by_xpath("/html/body/div[1]/section/div/div/div/p/span")
actions=ActionChains(driver)
actions.context_click(right_click).perform()




# ===========================================================
# Selenium with Python Tutorial 20-Handle Mouse Actions | Drag and Drop

from selenium import webdriver
from selenium.webdriver import ActionChains

driver=webdriver.Chrome(executable_path="chromedriver.exe")
driver.get("http://dhtmlgoodies.com/scripts/drag-drop-custom/demo-drag-drop-3.html")
driver.maximize_window()

source_element=driver.find_element_by_xpath("//*[@id='box6']")
target_element=driver.find_element_by_xpath("//*[@id='box106']")
source_element1=driver.find_element_by_xpath("//*[@id='box7']")
target_element1=driver.find_element_by_xpath("//*[@id='box107']")

actions=ActionChains(driver)
actions.drag_and_drop(source_element,target_element).perform()
actions.drag_and_drop(source_element1,target_element1).perform()




# ==============================================================
# Selenium with Python Tutorial 21- How to upload Files

from selenium import webdriver

driver=webdriver.Chrome(executable_path="chromedriver.exe")
driver.get("http://testautomationpractice.blogspot.com/")
driver.maximize_window()

driver.switch_to.frame(0)
driver.find_element_by_id("RESULT_FileUpload-10").send_keys("C:/Users/aslam/Pictures/Saved Pictures/hello2.gif")
driver.find_element_by_id("RESULT_FileUpload-10").send_keys("C:/Users/aslam/Pictures/Wallpapers/8972.jpg")



# ===========================================================
# Selenium with Python Tutorial 22- How to Download Files using Chrome Browser

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

chromeoptions=Options()
chromeoptions.add_experimental_option("prefs",{"download_default_directory": "C:/Users/aslam/Desktop/Govt_Job"})

driver=webdriver.Chrome(executable_path="chromedriver.exe",options=chromeoptions)
driver.get("http://demo.automationtesting.in/FileDownload.html")
driver.maximize_window()

driver.find_element_by_id("textbox").send_keys("Hello Faruk, Whats up bro!")
driver.find_element_by_id("createTxt").click()
driver.find_element_by_id("link-to-download").click()

driver.find_element_by_id("pdfbox").send_keys("hello fuck you madarchod!")
driver.find_element_by_id("createPdf").click()
driver.find_element_by_id("pdf-link-to-download").click()





# ===========================================================
# Selenium with Python Tutorial 23- How to Download Files using Firefox Browser

from selenium import webdriver

fp=webdriver.FirefoxProfile()
fp.set_preference("browser.helperApps.neverAsk.saveToDisk","text/plain,application/pdf")
fp.set_preference("browser.download.manager.showWhenStarting",False)
fp.set_preference("browser.download.dir","C:/Users/aslam/Desktop/Govt_Job")
fp.set_preference("browser.download.folderList",2)
fp.set_preference("pdfjs.disabled",True)

driver=webdriver.Firefox(executable_path="geckodriver.exe",firefox_profile=fp)
driver.get("http://demo.automationtesting.in/FileDownload.html")
driver.maximize_window()

driver.find_element_by_id("textbox").send_keys("Hello Faruk, Whats up bro!")
driver.find_element_by_id("createTxt").click()
driver.find_element_by_id("link-to-download").click()

driver.find_element_by_id("pdfbox").send_keys("hello fuck you madarchod!")
driver.find_element_by_id("createPdf").click()
driver.find_element_by_id("pdf-link-to-download").click()



# ===========================================================
# Selenium with Python Tutorial 24-How to read data from MS-Excel using OpenPyXL | Data Driven Testing

import openpyxl

path="S:/Python Programming/Amulyas Academy/hello.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active # workbook.get_sheet_by_name("sheet1")

rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="  ")
    print()

"""

# ===========================================================
# Selenium with Python Tutorial 25-How to write data into Excel using OpenPyXL | Data Driven Testing

import openpyxl

path="hello1.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value="welcome"

workbook.save(path)






